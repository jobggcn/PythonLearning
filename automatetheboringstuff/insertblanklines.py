#Excercise 2 Chapter 12 
#Blank Line Inserter 

import openpyxl, sys, os


def insert_blank_lines(file_name, insert_from, gap):
    #This implementation moves lower rows instead of copying them to memory and writing them again
    #I'm fairly certain that this is less memory intensive and maybe even a little faster
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.get_active_sheet()
    
    sheet_height = sheet.get_highest_row() 
    sheet_width = sheet.get_highest_column() 
    
    #Move cells below 'insert_from' by 'gap'
    for row_index in reversed(range(insert_from, sheet_height)): #iterates from bottom to top
        for column_index in range(sheet_width):
            #copy the cell to the cell x indices below given by gap
            sheet.cell(row = row_index + gap + 1, column = column_index + 1).value = \
            sheet.cell(row = row_index + 1, column = column_index + 1).value
    #Empty cells within gap
    for row_index in range(insert_from, insert_from + gap):
        for column_index in range(sheet_width):
            sheet.cell(row = row_index + 1, column = column_index + 1).value = ''

    workbook.save(file_name)

if len(sys.argv) == 4:
    if not ( os.path.isfile(sys.argv[1]) and sys.argv[1].endswith('xlsx') ):
        print('Invalid file path. File needs to be .xlsx file')
    else:
        insert_blank_lines(sys.argv[1], int(sys.argv[2]), int(sys.argv[3]))
else:
    print('insertblanklines.py <file path> <starting row> <gap width>') 

