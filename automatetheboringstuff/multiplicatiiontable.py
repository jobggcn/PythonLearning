#Excercise 1 Chapter 12 
#Multiplication Table Maker 
# x 1 2 3 4 
# 1 1 2 3 4
# 2 2 4 6 8
# 3 3 6 9 12
# 4 4 8 1216

import openpyxl, sys 
from openpyxl.styles import Font, Style


def create_multiplication_table(size):
    workbook = openpyxl.Workbook()
    sheet = workbook.get_active_sheet()
    #Labels
    for i in range(1, size + 1): #fill first row 
        sheet.cell(row = 1, column = i + 1).value = i  
        for i in range(1, size + 1): #fill first column 
            sheet.cell(row = i + 1, column = 1).value = i 
    #Table 
    for h_index  in range(1, size + 1): #fill table 
        for v_index in range(1, size + 1): 
            sheet.cell(row = v_index + 1, column = h_index + 1).value = h_index * v_index
    #Styling
    boldFontObject = Font(bold = True) 
    boldStyleObject = Style(font = boldFontObject)
    # Set first row and column to bold  
    for cell in sheet.columns[0]:
        cell.style = boldStyleObject 
    for cell in sheet.rows[0]:
        cell.style = boldStyleObject 
    
    workbook.save('multiplicationtable{}.xlsx'.format(size))

if len(sys.argv) == 2:
    create_multiplication_table(int(sys.argv[1]))
else:
    print('python3 multiplicationtable.py <size>') 
