#Excercise 3 Chapter 12 
#Spreadsheet Cell Inverter

import openpyxl, os, sys  
import pprint

def find_accurate_column_number(sheet):
    '''
    Due to libre office bumping up column numbers to 1024, this function gives column number ignoring empty rightbound columns 
    Since the libre office dev team considers extending the file by 1024 empty columns as a feature, i'll just have to use this.
    Takes several seconds to execute too...
    '''
    sheet_height = sheet.get_highest_row()
    sheet_width_upper_bound = sheet.get_highest_column()
    for column_index in reversed(range(sheet_width_upper_bound)):
        for row_index in range(sheet_width_upper_bound):
            if not sheet.cell(row = row_index + 1, column = column_index + 1).value is None:
                return column_index + 1
    return 0 #Empty sheet



def rotate_sheet(file_name): 
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.get_active_sheet()
    sheet_height = sheet.get_highest_row() 
    sheet_width = find_accurate_column_number(sheet) 
    
    #Write sheet contents into cell buffer and clear sheet
    cell_buffer = []
    for row_index in range(sheet_height): 
        row = []
        for column_index in range(sheet_width + 1):
            row.append(sheet.cell(row = row_index + 1, column = column_index + 1).value)
            sheet.cell(row = row_index + 1, column = column_index + 1).value = '' 
        cell_buffer.append(row) 
    #Write cell buffer back into sheet inverting the row and column numbers
    for row_index in range(sheet_height):
        for column_index in range(sheet_width):
            sheet.cell(row = column_index + 1, column = row_index + 1).value = cell_buffer[row_index][column_index]
    workbook.save(file_name)

if len(sys.argv) == 2:
    if os.path.isfile(sys.argv[1]) and sys.argv[1].endswith('.xlsx'):
        rotate_sheet(sys.argv[1])
    else:
        print("Invalid File Path") 
else:
    print('python3 spreadsheetcellinverter.py <spreadsheet path>') 
