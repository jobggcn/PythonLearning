#Exercise 5 Chapter 12 
#Spreadsheet to Text Files
#Does the reverse of the Exercise 4 Script

import openpyxl, os, sys 
from openpyxl.cell import get_column_letter 
 

def write_text_files(sheet_path):
    workbook = openpyxl.load_workbook(sheet_path)
    sheet = workbook.get_active_sheet() 
    
    for column in sheet.columns: 
        values = [cell.value for cell in column] 
        if not all(x is None for x in values):
            textfile = open( sheet.title + get_column_letter(sheet.columns.index(column)+1) + '.txt', 'w')
            values = [value or '' for value in values] 
            for value in values: 
                textfile.write(value  + '\n') 
        textfile.close() 
        
if len(sys.argv) == 2: 
    if os.path.isfile(sys.argv[1]) and sys.argv[1].endswith('.xlsx'):
        write_text_files(sys.argv[1])
    else:
        print("Invalid File: File needs to be .xlsx") 
else:
    print('python3 xlsxtotxt.py <spreadsheet path.') 

